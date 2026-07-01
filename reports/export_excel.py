"""
导出 Excel 计算书
"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(ws, row, cols, values, fill_color="34495e"):
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=cols[0] + c - 1 if isinstance(cols, list) else c, value=v)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def _write_table(ws, start_row, headers, data_rows, start_col=1):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    _style_header(ws, start_row, list(range(start_col, start_col + len(headers))), headers)
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.border = thin_border
            cell.font = Font(name="Microsoft YaHei", size=10)


def _kv_section(ws, row, title, data):
    ws.cell(row=row, column=1, value=title).font = Font(name="Microsoft YaHei", bold=True, size=13)
    row += 1
    for k, v in data.items():
        ws.cell(row=row, column=1, value=str(k)).font = Font(name="Microsoft YaHei", bold=True, size=10)
        ws.cell(row=row, column=2, value=str(v)).font = Font(name="Microsoft YaHei", size=10)
        row += 1
    return row + 1


def export_to_excel(collected: dict, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "\u8bbe\u8ba1\u8ba1\u7b97\u4e66"
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="\u7ba1\u5f0f\u819c\u8bbe\u8ba1\u8ba1\u7b97\u4e66").font = Font(name="Microsoft YaHei", bold=True, size=16, color="2c3e50")
    row += 2
    # project info
    if "\u9879\u76ee\u57fa\u7840\u4fe1\u606f" in collected:
        pd = collected["\u9879\u76ee\u57fa\u7840\u4fe1\u606f"]
        ws.cell(row=row, column=1, value="\u4e00\u3001\u9879\u76ee\u57fa\u7840\u4fe1\u606f").font = Font(name="Microsoft YaHei", bold=True, size=13)
        row += 1
        fields = [("\u9879\u76ee\u540d\u79f0", pd.get("project_name", "")), ("\u9879\u76ee\u7f16\u53f7", pd.get("project_no", "")), ("\u8bbe\u8ba1\u4ea7\u6c34\u91cf", f"{pd.get('design_flow', '')} m\u00b3/d"), ("\u6c34\u578b", pd.get("water_type", "")), ("\u8bbe\u8ba1\u4eba\u5458", pd.get("designer", "")), ("\u8bbe\u8ba1\u65e5\u671f", pd.get("design_date", ""))]
        for label, val in fields:
            ws.cell(row=row, column=1, value=label).font = Font(name="Microsoft YaHei", bold=True, size=10)
            ws.cell(row=row, column=2, value=str(val)).font = Font(name="Microsoft YaHei", size=10)
            row += 1
        row += 1
    # membrane
    if "\u819c\u578b\u53f7" in collected:
        row = _kv_section(ws, row, "\u4e8c\u3001\u819c\u578b\u53f7\u53c2\u6570", collected["\u819c\u578b\u53f7"])
    # area calc
    if "\u819c\u9762\u79ef\u8ba1\u7b97" in collected:
        row = _kv_section(ws, row, "\u4e09\u3001\u819c\u9762\u79ef\u8ba1\u7b97", collected["\u819c\u9762\u79ef\u8ba1\u7b97"])
    # crossflow
    if "\u9519\u6d41\u5faa\u73af\u91cf" in collected:
        row = _kv_section(ws, row, "\u56db\u3001\u9519\u6d41\u5faa\u73af\u91cf\u8ba1\u7b97", collected["\u9519\u6d41\u5faa\u73af\u91cf"])
    # TMP
    if "TMP\u538b\u529b\u6821\u6838" in collected:
        rd = collected["TMP\u538b\u529b\u6821\u6838"]
        ws.cell(row=row, column=1, value="\u4e94\u3001TMP\u538b\u529b\u6821\u6838").font = Font(name="Microsoft YaHei", bold=True, size=13)
        row += 1
        for k, v in rd.items():
            if k == "warnings":
                v = "; ".join(v)
            ws.cell(row=row, column=1, value=str(k)).font = Font(name="Microsoft YaHei", bold=True, size=10)
            ws.cell(row=row, column=2, value=str(v)).font = Font(name="Microsoft YaHei", size=10)
            row += 1
        row += 1
    # pumps
    if "\u6cf5\u9009\u578b" in collected:
        rd = collected["\u6cf5\u9009\u578b"]
        ws.cell(row=row, column=1, value="\u516d\u3001\u6cf5\u9009\u578b").font = Font(name="Microsoft YaHei", bold=True, size=13)
        row += 1
        headers = ["\u6cf5\u540d\u79f0", "\u6d41\u91cf(m\u00b3/h)", "\u626c\u7a0b(m)", "\u8f74\u529f\u7387(kW)", "\u7535\u673a\u529f\u7387(kW)", "\u6750\u8d28\u5efa\u8bae"]
        rows = []
        for key in ["feed_pump", "crossflow_pump", "cleaning_pump"]:
            p = rd.get(key, {})
            rows.append([p.get("name", key), f"{p.get('flow', 0):.1f}", f"{p.get('head', 0):.0f}", f"{p.get('shaft_power_kw', 0):.2f}", f"{p.get('recommended_motor_kw', 0):.0f}", p.get("material", "")])
        _write_table(ws, row, headers, rows)
        row += len(rows) + 2
    # pipes
    if "\u7ba1\u5f84\u9009\u578b" in collected:
        rd = collected["\u7ba1\u5f84\u9009\u578b"]
        ws.cell(row=row, column=1, value="\u4e03\u3001\u7ba1\u5f84\u9009\u578b").font = Font(name="Microsoft YaHei", bold=True, size=13)
        row += 1
        pipes = rd.get("pipes", [])
        headers = ["\u7ba1\u8def\u7c7b\u578b", "\u6d41\u91cf(m\u00b3/h)", "\u63a8\u8350DN", "\u5185\u5f84(mm)", "\u6d41\u901f(m/s)", "\u8bc4\u4ef7"]
        rows = [[p["pipe_type"], f"{p['flow']:.1f}", p["recommended_dn"], p["inner_diameter_mm"], f"{p['actual_velocity']:.2f}", p["advice"]] for p in pipes]
        _write_table(ws, row, headers, rows)
        row += len(rows) + 2
    # CIP
    if "CIP\u6e05\u6d17" in collected:
        row = _kv_section(ws, row, "\u516b\u3001CIP\u6e05\u6d17\u8ba1\u7b97", collected["CIP\u6e05\u6d17"])
    # bubble point
    if "\u6ce1\u70b9\u68c0\u6d4b" in collected:
        rd = collected["\u6ce1\u70b9\u68c0\u6d4b"]
        ws.cell(row=row, column=1, value="\u4e5d\u3001\u6ce1\u70b9\u68c0\u6d4b\u6c47\u603b").font = Font(name="Microsoft YaHei", bold=True, size=13)
        row += 1
        summary = {k: v for k, v in rd.items() if k != "records"}
        for k, v in summary.items():
            ws.cell(row=row, column=1, value=str(k)).font = Font(name="Microsoft YaHei", bold=True, size=10)
            ws.cell(row=row, column=2, value=str(v)).font = Font(name="Microsoft YaHei", size=10)
            row += 1
        row += 1
    # BOM
    if "BOM\u6e05\u5355" in collected:
        rd = collected["BOM\u6e05\u5355"]
        ws.cell(row=row, column=1, value="\u5341\u3001BOM\u6e05\u5355").font = Font(name="Microsoft YaHei", bold=True, size=13)
        row += 1
        items = rd.get("items", [])
        headers = ["\u5e8f\u53f7", "\u540d\u79f0", "\u89c4\u683c", "\u6570\u91cf", "\u5355\u4f4d", "\u5355\u4ef7", "\u91d1\u989d", "\u5907\u6ce8"]
        rows = [[str(item.get(k, "")) for k in ["seq", "name", "spec", "quantity", "unit", "price", "amount", "remark"]] for item in items]
        _write_table(ws, row, headers, rows)
        row += len(rows) + 2
        ws.cell(row=row, column=1, value=f"\u603b\u4ef7: {rd.get('total', 0):,.2f} \u5143").font = Font(name="Microsoft YaHei", bold=True, size=12, color="e74c3c")
    # column width
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
